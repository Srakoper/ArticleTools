"""
Module for writing articles from SQL database to .xls[x] file.
Takes start date and end date as an input.
Fetches all data on articles from a given period, except ids from SQL tables, important text, tag relevance, tag similarity, number of tags.
Appends tags of an article in a form "tag1[, tag2, tag3...]" in a single .xls[x] cell.
"""

import sqlite3
import openpyxl
import re
from time import time as t

#database = input("Enter database name: ")
while True:
    try:
        begin = input("Articles from (date in YYYY-MM-DD format): ")
        assert re.search(r"20\d{2}-\d{2}-\d{2}", begin)
        break
    except AssertionError: print("Date not in required format. Enter a date in YYYY-MM-DD format.")
while True:
    try:
        end = input("Articles to (date in YYYY-M-D format, to >= from): ")
        assert re.search(r"20\d{2}-\d{2}-\d{2}", end)
        break
    except AssertionError: print("Date not in required format. Enter a date in YYYY-M-D format.")
"""
xls file
"""
wb = openpyxl.Workbook()
sheet = wb.active
sheet["A1"] = "ID"
#sheet["A1"].font.bold = True
sheet["B1"] = "Address"
# sheet.cell("B1").style.font.bold = True
sheet["C1"] = "Section"
# sheet.cell("C1").style.font.bold = True
sheet["D1"] = "Author"
# sheet.cell("D1").style.font.bold = True
sheet["E1"] = "Coauthors"
# sheet.cell("E1").style.font.bold = True
sheet["F1"] = "Time"
# sheet.cell("F1").style.font.bold = True
sheet["G1"] = "Title"
# sheet.cell("G1").style.font.bold = True
sheet["H1"] = "Label"
# sheet.cell("H1").style.font.bold = True
sheet["I1"] = "Lead"
# sheet.cell("I1").style.font.bold = True
sheet["J1"] = "Content"
# sheet.cell("J1").style.font.bold = True
sheet["K1"] = "Tags"
# sheet.cell("K1").style.font.bold = True
counter = 2
"""
Executing SQL queries.
"""
start = t()
connection = sqlite3.connect("articles.sqlite")
cursor = connection.cursor()
id_number = 0
for row in cursor.execute("""   SELECT * FROM Articles
                                LEFT OUTER JOIN Relations ON Articles.id = Relations.id_article
                                LEFT OUTER JOIN Tags ON Relations.id_tag = Tags.id
                                WHERE (Articles.time BETWEEN ? AND ?)
                                AND (section LIKE '%/dom%' OR section LIKE '%/vrt%')
                                ORDER BY Articles.time DESC""", (begin, end)):
    if row[1] != id_number:
        sheet["A" + str(counter)] = row[1]
        sheet["B" + str(counter)] = row[2]
        sheet["C" + str(counter)] = row[3]
        sheet["D" + str(counter)] = row[4]
        sheet["E" + str(counter)] = row[14]
        sheet["F" + str(counter)] = row[5]
        sheet["G" + str(counter)] = row[6]
        sheet["H" + str(counter)] = row[7]
        sheet["I" + str(counter)] = row[8]
        sheet["J" + str(counter)] = row[9]
        sheet["K" + str(counter)] = row[18]
        counter += 1
        id_number = row[1]
    else: sheet["K" + str(counter - 1)] = sheet["K" + str(counter - 1)].value + ", " + row[18]
while True:
    try:
        wb.save("C:\\Users\\dmihelic\\Desktop\\Tools\\export.xlsx")
        break
    except PermissionError: input("Please close export.xlsx and press any key. ")

print("Finished in %s seconds." % "{0:.3f}".format(t() - start))